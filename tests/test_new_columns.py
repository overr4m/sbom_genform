"""
Tests for the new report columns added in the last two increments:

Component columns
─────────────────
 • Тип пакета / тип компонента   (package_type    from PURL)
 • PURL / технический идентификатор компонента (purl)
 • Признак принадлежности к поверхности атаки  (attack_surface)
 • Признак выполнения функций безопасности      (security_function)
 • Принадлежность к контейнерному образу        (container_image)
 • Роль компонента в составе контейнерного образа (container_role)

Vulnerability columns
─────────────────────
 • Рекомендация / компенсирующая мера           (recommendation)
 • Статус допустимости в рассматриваемой конфигурации (acceptability_status)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict
from unittest.mock import patch

import pytest

from sbom_pipeline.vuln_merger import VulnFinding, merge_vulns_into_sbom, save_vuln_report
from sbom_pipeline.exporter import Exporter, _COMP_COLUMNS, _VULN_COLUMNS

# ---------------------------------------------------------------------------
# Helpers / shared fixtures
# ---------------------------------------------------------------------------

def _sbom(**overrides) -> Dict[str, Any]:
    """Minimal valid CycloneDX SBOM dict."""
    base: Dict[str, Any] = {
        "bomFormat": "CycloneDX",
        "specVersion": "1.5",
        "components": [],
    }
    base.update(overrides)
    return base


def _comp(
    name: str = "libfoo",
    version: str = "1.0.0",
    purl: str = "pkg:pypi/libfoo@1.0.0",
    properties: list | None = None,
) -> Dict[str, Any]:
    c: Dict[str, Any] = {"type": "library", "name": name, "version": version, "purl": purl}
    if properties is not None:
        c["properties"] = properties
    return c


def _vuln(
    cve_id: str = "CVE-2024-0001",
    scanner: str = "trivy",
    recommendation: str = "",
    acceptability_status: str = "",
) -> VulnFinding:
    return VulnFinding(
        cve_id=cve_id,
        component_name="libfoo",
        component_version="1.0.0",
        component_purl="pkg:pypi/libfoo@1.0.0",
        cvss_score=7.5,
        severity="HIGH",
        description="Test vulnerability",
        scanner=scanner,
        fixed_version="1.1.0",
        recommendation=recommendation,
        acceptability_status=acceptability_status,
    )


# ===========================================================================
# 1. _purl_type helper
# ===========================================================================

class TestPurlType:
    """pipeline._purl_type() extracts the ecosystem from a PURL string."""

    def _purl_type(self, purl: str) -> str:
        from sbom_pipeline.pipeline import _purl_type
        return _purl_type(purl)

    def test_pypi(self):
        assert self._purl_type("pkg:pypi/requests@2.31.0") == "pypi"

    def test_maven(self):
        assert self._purl_type("pkg:maven/org.springframework/spring-core@5.3.0") == "maven"

    def test_npm(self):
        assert self._purl_type("pkg:npm/%40angular/core@15.0.0") == "npm"

    def test_apk(self):
        assert self._purl_type("pkg:apk/alpine/openssl@1.1.1k-r0") == "apk"

    def test_deb(self):
        assert self._purl_type("pkg:deb/debian/curl@7.64.0-4") == "deb"

    def test_empty(self):
        assert self._purl_type("") == ""

    def test_non_purl(self):
        assert self._purl_type("not-a-purl") == ""


# ===========================================================================
# 2. _find_prop helper
# ===========================================================================

class TestFindProp:
    """pipeline._find_prop() returns the first matching key value."""

    def _find_prop(self, props, keys):
        from sbom_pipeline.pipeline import _find_prop
        return _find_prop(props, keys)

    def test_first_key_wins(self):
        props = {"attack-surface": "yes", "attackSurface": "no"}
        assert self._find_prop(props, ("attack-surface", "attackSurface")) == "yes"

    def test_fallback_key(self):
        props = {"attackSurface": "yes"}
        assert self._find_prop(props, ("attack-surface", "attackSurface")) == "yes"

    def test_no_match_returns_empty(self):
        props = {"other": "value"}
        assert self._find_prop(props, ("attack-surface", "attackSurface")) == ""

    def test_empty_props(self):
        assert self._find_prop({}, ("attack-surface",)) == ""


# ===========================================================================
# 3. _extract_dependencies — component columns from SBOM
# ===========================================================================

class TestExtractDependencies:
    """pipeline._extract_dependencies() populates all new Dependency attributes."""

    def _extract(self, sbom: Dict[str, Any]) -> list:
        from sbom_pipeline.pipeline import _extract_dependencies
        # Patch Dependency._process_purl to skip HTTP calls
        with patch("sbom_pipeline.dependency.Dependency._process_purl"):
            return _extract_dependencies(sbom, "/fake/sbom.json")

    # -----------------------------------------------------------------------
    # package_type
    # -----------------------------------------------------------------------

    def test_package_type_pypi(self):
        sbom = _sbom(components=[_comp(purl="pkg:pypi/flask@3.0.0")])
        deps = self._extract(sbom)
        assert deps[0].package_type == "pypi"

    def test_package_type_maven(self):
        sbom = _sbom(components=[_comp(purl="pkg:maven/com.google.guava/guava@32.0.0")])
        deps = self._extract(sbom)
        assert deps[0].package_type == "maven"

    def test_package_type_apk(self):
        sbom = _sbom(components=[_comp(purl="pkg:apk/alpine/openssl@1.1.1k-r0")])
        deps = self._extract(sbom)
        assert deps[0].package_type == "apk"

    def test_package_type_empty_when_no_purl(self):
        comp = {"type": "library", "name": "unknown", "version": "0.0.1"}
        sbom = _sbom(components=[comp])
        deps = self._extract(sbom)
        assert deps[0].package_type == ""

    # -----------------------------------------------------------------------
    # purl
    # -----------------------------------------------------------------------

    def test_purl_is_preserved(self):
        purl = "pkg:npm/lodash@4.17.21"
        sbom = _sbom(components=[_comp(purl=purl)])
        deps = self._extract(sbom)
        assert deps[0].purl == purl

    # -----------------------------------------------------------------------
    # attack_surface
    # -----------------------------------------------------------------------

    @pytest.mark.parametrize("prop_name", [
        "attack-surface",
        "attack_surface",
        "attackSurface",
        "isAttackSurface",
    ])
    def test_attack_surface_recognised(self, prop_name: str):
        props = [{"name": prop_name, "value": "yes"}]
        sbom = _sbom(components=[_comp(properties=props)])
        deps = self._extract(sbom)
        assert deps[0].attack_surface == "yes"

    def test_attack_surface_empty_when_absent(self):
        sbom = _sbom(components=[_comp(properties=[])])
        deps = self._extract(sbom)
        assert deps[0].attack_surface == ""

    # -----------------------------------------------------------------------
    # security_function
    # -----------------------------------------------------------------------

    @pytest.mark.parametrize("prop_name", [
        "security-function",
        "security_function",
        "securityFunction",
        "isSecurityFunction",
    ])
    def test_security_function_recognised(self, prop_name: str):
        props = [{"name": prop_name, "value": "crypto"}]
        sbom = _sbom(components=[_comp(properties=props)])
        deps = self._extract(sbom)
        assert deps[0].security_function == "crypto"

    def test_security_function_empty_when_absent(self):
        sbom = _sbom(components=[_comp()])
        deps = self._extract(sbom)
        assert deps[0].security_function == ""

    # -----------------------------------------------------------------------
    # container_image
    # -----------------------------------------------------------------------

    def test_container_image_from_metadata(self):
        sbom = _sbom(
            metadata={"component": {"type": "container", "name": "my-image:latest"}},
            components=[_comp()],
        )
        deps = self._extract(sbom)
        assert deps[0].container_image == "my-image:latest"

    def test_container_image_empty_for_non_container_metadata(self):
        sbom = _sbom(
            metadata={"component": {"type": "application", "name": "my-app"}},
            components=[_comp()],
        )
        deps = self._extract(sbom)
        assert deps[0].container_image == ""

    def test_container_image_empty_when_no_metadata(self):
        sbom = _sbom(components=[_comp()])
        deps = self._extract(sbom)
        assert deps[0].container_image == ""

    def test_container_image_from_component_property(self):
        """container_image property set by Clair enrichment takes priority over metadata."""
        sbom = _sbom(
            metadata={"component": {"type": "container", "name": "old-image:1.0"}},
            components=[_comp(properties=[{"name": "container_image", "value": "new-image:2.0"}])],
        )
        deps = self._extract(sbom)
        assert deps[0].container_image == "new-image:2.0"

    def test_container_image_property_fallback_to_metadata_when_prop_absent(self):
        """Without per-component property, still falls back to metadata container name."""
        sbom = _sbom(
            metadata={"component": {"type": "container", "name": "meta-image:1.0"}},
            components=[_comp()],
        )
        deps = self._extract(sbom)
        assert deps[0].container_image == "meta-image:1.0"
    # -----------------------------------------------------------------------
    # container_role
    # -----------------------------------------------------------------------

    @pytest.mark.parametrize("prop_name", [
        "container-role",
        "container_role",
        "containerRole",
        "cdx:docker:layer",
        "layer",
    ])
    def test_container_role_recognised(self, prop_name: str):
        props = [{"name": prop_name, "value": "os-packages"}]
        sbom = _sbom(components=[_comp(properties=props)])
        deps = self._extract(sbom)
        assert deps[0].container_role == "os-packages"

    def test_container_role_empty_when_absent(self):
        sbom = _sbom(components=[_comp()])
        deps = self._extract(sbom)
        assert deps[0].container_role == ""

    # -----------------------------------------------------------------------
    # Non-library components are skipped
    # -----------------------------------------------------------------------

    def test_non_library_components_skipped(self):
        sbom = _sbom(components=[
            {"type": "application", "name": "app", "version": "1.0", "purl": "pkg:pypi/app@1.0"},
            _comp(name="libfoo"),
        ])
        deps = self._extract(sbom)
        assert len(deps) == 1
        assert deps[0].name == "libfoo"

    # -----------------------------------------------------------------------
    # Image SBOM fixture (all fields together)
    # -----------------------------------------------------------------------

    def test_image_sbom_fixture(self):
        fixture = Path(__file__).parent / "fixtures" / "sbom" / "images" / "sbom-image-sample.json"
        sbom = json.loads(fixture.read_text())
        deps = self._extract(sbom)
        assert len(deps) == 3
        # All components come from an apk PURL
        for dep in deps:
            assert dep.package_type == "apk"
        # Container image name from metadata
        for dep in deps:
            assert dep.container_image == "sample-backend-image"


# ===========================================================================
# 4. Scanner parsers — VulnFinding.recommendation & acceptability_status
# ===========================================================================

class TestTrivyRecommendation:
    """trivy._parse() fills recommendation and acceptability_status."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.trivy import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p, "trivy")
        p.unlink(missing_ok=True)
        return findings

    def _trivy_result(self, **vuln_overrides) -> dict:
        vuln = {
            "VulnerabilityID": "CVE-2024-1111",
            "PkgName": "libssl",
            "InstalledVersion": "1.1.1k",
            "PkgRef": "pkg:apk/libssl@1.1.1k",
            "Severity": "HIGH",
            "Title": "OpenSSL heap overflow",
            "FixedVersion": "1.1.1n",
            "CVSS": {},
        }
        vuln.update(vuln_overrides)
        return {"Results": [{"Vulnerabilities": [vuln]}]}

    def test_recommendation_from_primary_url(self):
        data = self._trivy_result(PrimaryURL="https://nvd.nist.gov/vuln/detail/CVE-2024-1111")
        findings = self._parse(data)
        assert findings[0].recommendation == "https://nvd.nist.gov/vuln/detail/CVE-2024-1111"

    def test_recommendation_fallback_to_fixed_version(self):
        data = self._trivy_result()  # no PrimaryURL
        findings = self._parse(data)
        assert findings[0].recommendation == "Обновить до версии 1.1.1n"

    def test_recommendation_empty_when_no_fixed_no_url(self):
        data = self._trivy_result(FixedVersion="")
        findings = self._parse(data)
        assert findings[0].recommendation == ""

    def test_acceptability_status_populated(self):
        data = self._trivy_result(Status="fixed")
        findings = self._parse(data)
        assert findings[0].acceptability_status == "Исправлено"

    def test_acceptability_status_empty_when_absent(self):
        data = self._trivy_result()
        findings = self._parse(data)
        assert findings[0].acceptability_status == ""

    def test_acceptability_status_will_not_fix(self):
        data = self._trivy_result(Status="will_not_fix")
        findings = self._parse(data)
        assert findings[0].acceptability_status == "Исправление не планируется"

    def test_acceptability_status_unknown_value_passed_through(self):
        data = self._trivy_result(Status="some_future_status")
        findings = self._parse(data)
        assert findings[0].acceptability_status == "some_future_status"


class TestClairRecommendation:
    """clair._parse() fills recommendation from Links."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.clair import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p)
        p.unlink(missing_ok=True)
        return findings

    def _clair_data(self, links: str = "") -> dict:
        """Build a minimal current-format Clair report JSON."""
        return {
            "manifest_hash": "sha256:test",
            "packages": {
                "1": {"id": "1", "name": "curl", "version": "7.64.0", "kind": "binary"}
            },
            "vulnerabilities": {
                "v1": {
                    "name": "CVE-2024-2222",
                    "description": "URL confusion in curl",
                    "severity": "high",
                    "normalized_severity": "High",
                    "links": links,
                    "fixed_in_version": "8.0.0",
                    "package": {},
                }
            },
            "package_vulnerabilities": {"1": ["v1"]},
            "environments": {},
            "distributions": {},
            "enrichments": {},
        }

    def test_recommendation_first_link(self):
        findings = self._parse(self._clair_data(links="https://cve.mitre.org/cgi-bin/cvename.cgi?name=CVE-2024-2222"))
        assert findings[0].recommendation == "https://cve.mitre.org/cgi-bin/cvename.cgi?name=CVE-2024-2222"

    def test_recommendation_empty_when_no_links(self):
        findings = self._parse(self._clair_data(links=""))
        assert findings[0].recommendation == ""

    def test_recommendation_empty_when_links_absent(self):
        data = {
            "manifest_hash": "sha256:test",
            "packages": {
                "1": {"id": "1", "name": "curl", "version": "7.64.0", "kind": "binary"}
            },
            "vulnerabilities": {
                "v1": {
                    "name": "CVE-2024-2222",
                    "description": "desc",
                    "severity": "high",
                    "normalized_severity": "High",
                    "fixed_in_version": "",
                    "package": {},
                }
            },
            "package_vulnerabilities": {"1": ["v1"]},
            "environments": {},
            "distributions": {},
            "enrichments": {},
        }
        findings = self._parse(data)
        assert findings[0].recommendation == ""


class TestDepcheckRecommendation:
    """depcheck._parse() fills recommendation from notes / references."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.depcheck import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p)
        p.unlink(missing_ok=True)
        return findings

    def _depcheck_data(self, notes: str = "", refs: list | None = None) -> dict:
        vuln: Dict[str, Any] = {
            "name": "CVE-2024-3333",
            "severity": "MEDIUM",
            "description": "Outdated library",
            "cvssv3": {"baseScore": 5.5},
        }
        if notes:
            vuln["notes"] = notes
        if refs is not None:
            vuln["references"] = refs
        return {
            "dependencies": [{
                "fileName": "log4j-1.2.jar",
                "packages": [{"id": "pkg:maven/log4j/log4j@1.2.17"}],
                "vulnerabilities": [vuln],
            }]
        }

    def test_recommendation_from_notes(self):
        findings = self._parse(self._depcheck_data(notes="Upgrade to log4j 2.x"))
        assert findings[0].recommendation == "Upgrade to log4j 2.x"

    def test_recommendation_from_references_url(self):
        refs = [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2024-3333", "name": "NVD"}]
        findings = self._parse(self._depcheck_data(refs=refs))
        assert findings[0].recommendation == "https://nvd.nist.gov/vuln/detail/CVE-2024-3333"

    def test_notes_takes_precedence_over_references(self):
        refs = [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2024-3333"}]
        findings = self._parse(self._depcheck_data(notes="Upgrade now", refs=refs))
        assert findings[0].recommendation == "Upgrade now"

    def test_recommendation_empty_when_no_notes_no_refs(self):
        findings = self._parse(self._depcheck_data())
        assert findings[0].recommendation == ""

    def test_recommendation_empty_ref_without_url(self):
        refs = [{"name": "NVD"}]  # no url key
        findings = self._parse(self._depcheck_data(refs=refs))
        assert findings[0].recommendation == ""


# ===========================================================================
# 4b. dependency-check — GHSA → CVE id extraction
# ===========================================================================

class TestDepcheckCveId:
    """depcheck._parse() resolves GHSA names to CVE IDs via NVD references."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.depcheck import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p)
        p.unlink(missing_ok=True)
        return findings

    def _depcheck_data(self, name: str, refs: list | None = None) -> dict:
        vuln: dict = {
            "name": name,
            "severity": "MEDIUM",
            "description": "test",
            "cvssv3": {"baseScore": 5.0},
        }
        if refs is not None:
            vuln["references"] = refs
        return {
            "dependencies": [{
                "fileName": "brace-expansion-3.0.0.tgz",
                "packages": [{"id": "pkg:npm/brace-expansion@3.0.0"}],
                "vulnerabilities": [vuln],
            }]
        }

    def test_ghsa_replaced_by_cve_from_nvd_url(self):
        refs = [
            {"url": "https://github.com/advisories/GHSA-f886-m6hf-6m8v"},
            {"url": "https://nvd.nist.gov/vuln/detail/CVE-2026-33750"},
        ]
        findings = self._parse(self._depcheck_data("GHSA-f886-m6hf-6m8v", refs))
        assert findings[0].cve_id == "CVE-2026-33750"

    def test_plain_cve_name_unchanged(self):
        refs = [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2023-1234"}]
        findings = self._parse(self._depcheck_data("CVE-2023-1234", refs))
        assert findings[0].cve_id == "CVE-2023-1234"

    def test_ghsa_kept_when_no_nvd_ref(self):
        refs = [
            {"url": "https://github.com/advisories/GHSA-f886-m6hf-6m8v"},
            {"url": "https://github.com/owner/repo/pull/96"},
        ]
        findings = self._parse(self._depcheck_data("GHSA-f886-m6hf-6m8v", refs))
        assert findings[0].cve_id == "GHSA-f886-m6hf-6m8v"

    def test_ghsa_kept_when_no_refs(self):
        findings = self._parse(self._depcheck_data("GHSA-f886-m6hf-6m8v"))
        assert findings[0].cve_id == "GHSA-f886-m6hf-6m8v"

    def test_cve_id_uppercased(self):
        refs = [{"url": "https://nvd.nist.gov/vuln/detail/cve-2026-33750"}]
        findings = self._parse(self._depcheck_data("GHSA-f886-m6hf-6m8v", refs))
        assert findings[0].cve_id == "CVE-2026-33750"


class TestDepcheckFixedVersion:
    """depcheck._parse() extracts fixed_version from vulnerableSoftware CPE ranges."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.depcheck import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p)
        p.unlink(missing_ok=True)
        return findings

    def _depcheck_data(self, vulnerable_software: list | None = None) -> dict:
        vuln: dict = {
            "name": "CVE-2024-1234",
            "severity": "HIGH",
            "description": "test",
            "cvssv3": {"baseScore": 7.5},
        }
        if vulnerable_software is not None:
            vuln["vulnerableSoftware"] = vulnerable_software
        return {
            "dependencies": [{
                "fileName": "brace-expansion-3.0.0.tgz",
                "packages": [{"id": "pkg:npm/brace-expansion@3.0.0"}],
                "vulnerabilities": [vuln],
            }]
        }

    def test_exclusive_upper_bound_extracted(self):
        # CPE as seen in the real report: \>\=2.0.0\<2.0.3
        sw = [{"software": {"id": r"cpe:2.3:a:*:brace-expansion:\>\=2.0.0\<2.0.3:*:*:*:*:*:*:*"}}]
        findings = self._parse(self._depcheck_data(sw))
        assert findings[0].fixed_version == "2.0.3"

    def test_inclusive_upper_bound_not_extracted(self):
        # <=2.0.1 — cannot determine the fix version
        sw = [{"software": {"id": r"cpe:2.3:a:*:brace-expansion:\>\=2.0.0\<\=2.0.1:*:*:*:*:*:*:*"}}]
        findings = self._parse(self._depcheck_data(sw))
        assert findings[0].fixed_version == ""

    def test_no_vulnerable_software_returns_empty(self):
        findings = self._parse(self._depcheck_data())
        assert findings[0].fixed_version == ""

    def test_empty_vulnerable_software_returns_empty(self):
        findings = self._parse(self._depcheck_data([]))
        assert findings[0].fixed_version == ""

    def test_real_report_cpe_example(self):
        # Mirrors GHSA-f886-m6hf-6m8v entry from the sample report
        sw = [{"software": {"id": r"cpe:2.3:a:*:brace-expansion:\>\=2.0.0\<2.0.3:*:*:*:*:*:*:*"}}]
        findings = self._parse(self._depcheck_data(sw))
        assert findings[0].fixed_version == "2.0.3"


class TestDepcheckAcceptabilityStatus:
    """depcheck._parse() sets acceptability_status from the 'unscored' flag."""

    def _parse(self, data: dict):
        import tempfile, json as _json
        from sbom_pipeline.scanner.depcheck import _parse
        with tempfile.NamedTemporaryFile(suffix=".json", mode="w", delete=False) as f:
            _json.dump(data, f)
            p = Path(f.name)
        findings = _parse(p)
        p.unlink(missing_ok=True)
        return findings

    def _depcheck_data(self, unscored: str | None = None) -> dict:
        vuln: dict = {
            "name": "CVE-2024-1234",
            "severity": "HIGH",
            "description": "test",
            "cvssv3": {"baseScore": 7.5},
        }
        if unscored is not None:
            vuln["unscored"] = unscored
        return {
            "dependencies": [{
                "fileName": "lib.tgz",
                "packages": [{"id": "pkg:npm/lib@1.0.0"}],
                "vulnerabilities": [vuln],
            }]
        }

    def test_unscored_true_sets_status(self):
        findings = self._parse(self._depcheck_data(unscored="true"))
        assert findings[0].acceptability_status == "Оценка не присвоена (advisory)"

    def test_unscored_absent_gives_empty(self):
        findings = self._parse(self._depcheck_data())
        assert findings[0].acceptability_status == ""

    def test_unscored_false_string_gives_empty(self):
        findings = self._parse(self._depcheck_data(unscored="false"))
        assert findings[0].acceptability_status == ""

    def test_real_report_ghsa_is_unscored(self):
        # All GHSA/NPM entries in the sample report carry "unscored": "true"
        data = {
            "dependencies": [{
                "fileName": "brace-expansion-3.0.0.tgz",
                "packages": [{"id": "pkg:npm/brace-expansion@3.0.0"}],
                "vulnerabilities": [{
                    "source": "NPM",
                    "name": "GHSA-f886-m6hf-6m8v",
                    "unscored": "true",
                    "severity": "moderate",
                    "cvssv3": {"baseScore": 6.5},
                    "description": "test",
                }],
            }]
        }
        findings = self._parse(data)
        assert findings[0].acceptability_status == "Оценка не присвоена (advisory)"


# ===========================================================================
# 5. VulnFinding dataclass — new fields present and default correctly
# ===========================================================================

class TestVulnFindingNewFields:
    def test_recommendation_default_empty(self):
        v = _vuln()
        assert v.recommendation == ""

    def test_acceptability_status_default_empty(self):
        v = _vuln()
        assert v.acceptability_status == ""

    def test_fields_set_correctly(self):
        v = _vuln(recommendation="Upgrade to 2.0", acceptability_status="acceptable")
        assert v.recommendation == "Upgrade to 2.0"
        assert v.acceptability_status == "acceptable"


# ===========================================================================
# 6. Exporter — _comp_rows() and _vuln_rows()
# ===========================================================================

class _FakeDep:
    """Minimal dependency object with all new attributes."""
    def __init__(self, **kwargs):
        self.name = kwargs.get("name", "libfoo")
        self.version = kwargs.get("version", "1.0.0")
        self.purl = kwargs.get("purl", "pkg:pypi/libfoo@1.0.0")
        self.package_type = kwargs.get("package_type", "pypi")
        self.srcLangs = kwargs.get("srcLangs", ["Python"])
        self.attack_surface = kwargs.get("attack_surface", "")
        self.security_function = kwargs.get("security_function", "")
        self.container_image = kwargs.get("container_image", "")
        self.container_role = kwargs.get("container_role", "")
        self.source = kwargs.get("source", "https://pypi.org/project/libfoo/1.0.0/")
        self.depType = kwargs.get("depType", [])


class TestExporterCompRows:
    """Exporter._comp_rows() maps all new component columns."""

    def _rows(self, **dep_kwargs) -> list:
        dep = _FakeDep(**dep_kwargs)
        exporter = Exporter([dep], vulns=[])
        return exporter._comp_rows()

    def test_package_type_column(self):
        rows = self._rows(package_type="maven")
        assert rows[0]["Тип пакета / тип компонента"] == "maven"

    def test_purl_column(self):
        rows = self._rows(purl="pkg:npm/lodash@4.17.21")
        assert rows[0]["PURL / технический идентификатор компонента"] == "pkg:npm/lodash@4.17.21"

    def test_attack_surface_column(self):
        rows = self._rows(attack_surface="yes")
        assert rows[0]["Признак принадлежности к поверхности атаки"] == "yes"

    def test_attack_surface_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Признак принадлежности к поверхности атаки"] == ""

    def test_security_function_column(self):
        rows = self._rows(security_function="tls")
        assert rows[0]["Признак выполнения функций безопасности"] == "tls"

    def test_security_function_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Признак выполнения функций безопасности"] == ""

    def test_container_image_column(self):
        rows = self._rows(container_image="my-image:1.0")
        assert rows[0]["Принадлежность к контейнерному образу"] == "my-image:1.0"

    def test_container_image_empty_for_git_sbom(self):
        rows = self._rows(container_image="")
        assert rows[0]["Принадлежность к контейнерному образу"] == ""

    def test_container_role_column(self):
        rows = self._rows(container_role="os-packages")
        assert rows[0]["Роль компонента в составе контейнерного образа"] == "os-packages"

    def test_container_role_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Роль компонента в составе контейнерного образа"] == ""

    def test_web_address_from_source(self):
        rows = self._rows(source="https://pypi.org/project/libfoo/")
        assert rows[0]["Адрес веб-ресурса"] == "https://pypi.org/project/libfoo/"

    def test_all_expected_columns_present(self):
        rows = self._rows()
        for col in _COMP_COLUMNS:
            assert col in rows[0], f"Missing column: {col}"

    def test_row_number_increments(self):
        dep1 = _FakeDep(name="a")
        dep2 = _FakeDep(name="b")
        exporter = Exporter([dep1, dep2], vulns=[])
        rows = exporter._comp_rows()
        assert rows[0]["№ п/п"] == 1
        assert rows[1]["№ п/п"] == 2

    def test_old_attack_surface_column_no_longer_present(self):
        """The merged column must NOT exist anymore."""
        rows = self._rows()
        assert "Принадлежность к поверхности атаки / функциям безопасности" not in rows[0]


class TestExporterVulnRows:
    """Exporter._vuln_rows() maps all new vulnerability columns."""

    def _rows(self, **vuln_kwargs) -> list:
        v = _vuln(**vuln_kwargs)
        exporter = Exporter([], vulns=[v])
        return exporter._vuln_rows()

    def test_recommendation_column_present(self):
        rows = self._rows(recommendation="Upgrade to 2.0")
        assert rows[0]["Рекомендация / компенсирующая мера"] == "Upgrade to 2.0"

    def test_recommendation_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Рекомендация / компенсирующая мера"] == ""

    def test_acceptability_status_column_present(self):
        rows = self._rows(acceptability_status="will_not_fix")
        assert rows[0]["Статус допустимости в рассматриваемой конфигурации"] == "will_not_fix"

    def test_acceptability_status_empty_by_default(self):
        rows = self._rows()
        assert rows[0]["Статус допустимости в рассматриваемой конфигурации"] == ""

    def test_all_expected_columns_present(self):
        rows = self._rows()
        for col in _VULN_COLUMNS:
            assert col in rows[0], f"Missing column: {col}"

    def test_existing_columns_still_present(self):
        rows = self._rows()
        assert rows[0]["CVE / ID"] == "CVE-2024-0001"
        assert rows[0]["Критичность"] == "HIGH"
        assert rows[0]["Исправлено в версии"] == "1.1.0"


# ===========================================================================
# 7. End-to-end: Excel export contains all new column headers
# ===========================================================================

class TestExcelExportHeaders:
    """The xlsx file produced by exportToExcel includes every new column."""

    def test_all_comp_columns_in_excel(self, tmp_path):
        import openpyxl

        dep = _FakeDep(
            package_type="pypi",
            purl="pkg:pypi/libfoo@1.0.0",
            attack_surface="yes",
            security_function="tls",
            container_image="my-image:1.0",
            container_role="app-layer",
        )
        out = str(tmp_path / "report.xlsx")
        Exporter([dep], vulns=[]).exportToExcel(out)

        wb = openpyxl.load_workbook(out)
        headers = [ws.cell(1, c).value for ws in wb.worksheets for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]

        for col in _COMP_COLUMNS:
            assert col in headers, f"Missing in Excel: {col}"

    def test_all_vuln_columns_in_excel(self, tmp_path):
        import openpyxl

        v = _vuln(recommendation="Upgrade", acceptability_status="fixed")
        out = str(tmp_path / "report.xlsx")
        Exporter([_FakeDep()], vulns=[v]).exportToExcel(out)

        wb = openpyxl.load_workbook(out)
        vuln_ws = wb["Уязвимости"]
        headers = [vuln_ws.cell(1, c).value for c in range(1, vuln_ws.max_column + 1)]

        for col in _VULN_COLUMNS:
            assert col in headers, f"Missing in Excel vuln sheet: {col}"

    def test_vuln_data_row_contains_new_fields(self, tmp_path):
        import openpyxl

        v = _vuln(recommendation="Обновить до 2.0", acceptability_status="end_of_life")
        out = str(tmp_path / "report.xlsx")
        Exporter([_FakeDep()], vulns=[v]).exportToExcel(out)

        wb = openpyxl.load_workbook(out)
        ws = wb["Уязвимости"]
        # Row 1 = header, row 2 = first data row
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        rec_col = headers["Рекомендация / компенсирующая мера"]
        status_col = headers["Статус допустимости в рассматриваемой конфигурации"]
        assert ws.cell(2, rec_col).value == "Обновить до 2.0"
        assert ws.cell(2, status_col).value == "end_of_life"


# ===========================================================================
# 8. merge_vulns_into_sbom — acceptability_status written to SBOM properties
# ===========================================================================

def _minimal_sbom_with_comp() -> dict:
    return {
        "bomFormat": "CycloneDX",
        "specVersion": "1.5",
        "components": [
            {
                "type": "library",
                "name": "libfoo",
                "version": "1.0.0",
                "purl": "pkg:pypi/libfoo@1.0.0",
                "bom-ref": "pkg:pypi/libfoo@1.0.0",
            }
        ],
    }


class TestMergeVulnsAcceptabilityStatus:
    """merge_vulns_into_sbom() writes acceptability_status as a CycloneDX property."""

    def test_acceptability_status_written_as_property(self):
        sbom = _minimal_sbom_with_comp()
        finding = _vuln(acceptability_status="Неприменимо")
        result = merge_vulns_into_sbom(sbom, [finding])
        vuln_entry = result["vulnerabilities"][0]
        props = vuln_entry.get("properties", [])
        status_props = [p for p in props if p["name"] == "acceptability_status"]
        assert len(status_props) == 1
        assert status_props[0]["value"] == "Неприменимо"

    def test_acceptability_status_absent_when_empty(self):
        sbom = _minimal_sbom_with_comp()
        finding = _vuln(acceptability_status="")
        result = merge_vulns_into_sbom(sbom, [finding])
        vuln_entry = result["vulnerabilities"][0]
        props = vuln_entry.get("properties", [])
        status_props = [p for p in props if p["name"] == "acceptability_status"]
        assert status_props == []

    def test_acceptability_status_coexists_with_bdu(self):
        sbom = _minimal_sbom_with_comp()
        finding = _vuln(acceptability_status="Неприменимо")
        finding.bdu_id = "BDU:2024-00001"
        result = merge_vulns_into_sbom(sbom, [finding], enable_bdu=False)
        # bdu_id was set manually, but enable_bdu=False so bdu lookup skipped;
        # acceptability_status should still appear as a property
        vuln_entry = result["vulnerabilities"][0]
        # With bdu_id set but no bdu lookup, we expect the property to be present
        props = vuln_entry.get("properties", [])
        status_props = [p for p in props if p["name"] == "acceptability_status"]
        assert len(status_props) == 1

    def test_fixed_version_and_acceptability_both_present(self):
        sbom = _minimal_sbom_with_comp()
        finding = _vuln(acceptability_status="Неприменимо")
        # _vuln() already sets fixed_version="1.1.0"
        result = merge_vulns_into_sbom(sbom, [finding])
        vuln_entry = result["vulnerabilities"][0]
        assert vuln_entry["recommendation"] == "Обновить до версии 1.1.0"
        props = vuln_entry.get("properties", [])
        assert any(p["value"] == "Неприменимо" for p in props)


# ===========================================================================
# 9. save_vuln_report — acceptability_status included in JSON output
# ===========================================================================

class TestSaveVulnReportAcceptabilityStatus:
    """save_vuln_report() includes acceptability_status in the saved JSON."""

    def test_acceptability_status_in_json(self, tmp_path):
        import json
        path = tmp_path / "vulns.json"
        finding = _vuln(acceptability_status="Неприменимо")
        save_vuln_report([finding], path)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert data[0]["acceptability_status"] == "Неприменимо"

    def test_acceptability_status_empty_string_in_json(self, tmp_path):
        import json
        path = tmp_path / "vulns.json"
        finding = _vuln(acceptability_status="")
        save_vuln_report([finding], path)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert data[0]["acceptability_status"] == ""

    def test_acceptability_status_key_always_present(self, tmp_path):
        import json
        path = tmp_path / "vulns.json"
        save_vuln_report([_vuln()], path)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert "acceptability_status" in data[0]
